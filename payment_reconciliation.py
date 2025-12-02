#!/usr/bin/env python3
"""
Payment Reconciliation System - Main Orchestrator
"""
import os
import sys
from datetime import datetime
import pandas as pd
from database import ReconciliationDB
from payment_advice_extractor import PaymentAdviceExtractor
from reconciliation_engine import ReconciliationEngine
from warsoft_client import WarsoftClient


def generate_excel_report(db):
    """Generate Excel report with MATCHED, UNMATCHED, and NOT_FOUND sheets"""
    print("\n📊 Generating Excel reconciliation report...")

    # Get today's date to filter only current run results
    today = datetime.now().strftime('%Y-%m-%d')

    # Get reconciliation results for today only
    results = db.get_all_reconciliation_results(date_filter=today)

    if not results:
        print(f"⚠️  No reconciliation data to report for {today}")
        return None

    # Convert to DataFrame
    df = pd.DataFrame([dict(row) for row in results])

    # Create Excel with multiple sheets
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'payment_reconciliation_{timestamp}.xlsx'

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Sheet 1: MATCHED
        matched_df = df[df['match_status'] == 'MATCHED'].copy()
        if not matched_df.empty:
            matched_df['category'] = 'MATCHED'
            matched_df.to_excel(writer, sheet_name='MATCHED', index=False)
            print(f"   ✅ MATCHED: {len(matched_df)} records")

        # Sheet 2: UNMATCHED (includes PARTIAL_MATCH)
        unmatched_df = df[df['match_status'].isin(['UNMATCHED', 'PARTIAL_MATCH'])].copy()
        if not unmatched_df.empty:
            unmatched_df['category'] = 'UNMATCHED'
            unmatched_df.to_excel(writer, sheet_name='UNMATCHED', index=False)
            print(f"   ⚠️  UNMATCHED: {len(unmatched_df)} records")

        # Sheet 3: NOT_FOUND
        not_found_df = df[df['match_status'] == 'NOT_FOUND'].copy()
        if not not_found_df.empty:
            not_found_df['category'] = 'NOT_FOUND'
            not_found_df.to_excel(writer, sheet_name='NOT_FOUND', index=False)
            print(f"   ❌ NOT_FOUND: {len(not_found_df)} records")

        # Sheet 4: ALL RESULTS
        df.to_excel(writer, sheet_name='ALL_RESULTS', index=False)

        # Sheet 5: SUMMARY
        summary_data = {
            'Category': [
                'Total Payments Processed',
                'MATCHED',
                'UNMATCHED (with discrepancies)',
                'NOT FOUND (invoice missing)',
                'Amount Mismatches',
                'Total Amount Difference'
            ],
            'Count/Value': [
                len(df),
                len(matched_df) if not matched_df.empty else 0,
                len(unmatched_df) if not unmatched_df.empty else 0,
                len(not_found_df) if not not_found_df.empty else 0,
                len(df[df['amount_match'] == False]),
                f"₹{df['amount_difference'].sum():.2f}"
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='SUMMARY', index=False)

        # Format columns for all sheets
        from openpyxl.styles import Font, PatternFill, Alignment

        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]

            # Header formatting
            if sheet_name != 'SUMMARY':
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Freeze top row
            worksheet.freeze_panes = 'A2'

    print(f"✅ Excel report generated: {filename}")
    return filename


def sync_invoices_from_warsoft(db, warsoft_client):
    """
    Sync unpaid invoices from Warsoft into database
    """
    print(f"\n📥 Fetching unpaid invoices from Warsoft...")

    # Fetch all unpaid invoices
    invoices = warsoft_client.fetch_all_unpaid_invoices()

    if not invoices:
        print("⚠️  No unpaid invoices found in Warsoft")
        return 0

    count = 0
    for invoice_raw in invoices:
        try:
            # Parse invoice to standardized format
            invoice = warsoft_client.parse_invoice(invoice_raw)
            db.insert_warsoft_invoice(invoice)
            count += 1
            print(
                f"   ✅ Synced: {invoice.get('invoice_number')} - ₹{invoice.get('total_amount', 0)} ({invoice.get('status')})")
        except Exception as e:
            print(f"   ⚠️  Error storing invoice {invoice_raw.get('invoiceNumber')}: {e}")

    print(f"✅ Synced {count} unpaid invoices from Warsoft")
    return count


def generate_no_invoice_report(db):
    """Generate separate Excel report for payment advices without invoice numbers"""
    print("\n📋 Checking for payment advices without invoice numbers...")

    # Get today's date to filter only current run results
    today = datetime.now().strftime('%Y-%m-%d')

    # Get payment advices without invoice numbers (today only)
    no_invoice_advices = db.get_payment_advices_without_invoice_numbers(date_filter=today)

    if not no_invoice_advices:
        print("   ✅ All payment advices have invoice numbers extracted!")
        return None

    print(f"   🔍 Found {len(no_invoice_advices)} payment advices without invoice numbers")

    # Convert to DataFrame
    no_invoice_df = pd.DataFrame([dict(row) for row in no_invoice_advices])

    # Check which columns are available
    available_columns = no_invoice_df.columns.tolist()
    print(f"   📋 Available columns: {', '.join(available_columns)}")

    # Select only available columns from our desired list (including invoice_number to show invalid extractions)
    desired_columns = ['id', 'invoice_number', 'pdf_filename', 'email_from', 'email_subject',
                       'email_date', 'payment_amount', 'net_payment_amount',
                       'bank_name', 'utr_number', 'status']

    columns_to_use = [col for col in desired_columns if col in available_columns]

    # Create simplified view with available columns
    report_df = no_invoice_df[columns_to_use].copy()

    # Rename columns for clarity (only the ones we have)
    column_mapping = {
        'id': 'ID',
        'invoice_number': 'Invalid Invoice Number',
        'pdf_filename': 'PDF Filename',
        'email_from': 'Email From',
        'email_subject': 'Email Subject',
        'email_date': 'Email Date',
        'payment_date': 'Payment Date',
        'payment_amount': 'Payment Amount',
        'net_payment_amount': 'Net Payment Amount',
        'bill_amount': 'Bill Amount',
        'tds_amount': 'TDS Amount',
        'bank_name': 'Bank Name',
        'utr_number': 'UTR Number',
        'status': 'Status'
    }

    report_df.columns = [column_mapping.get(col, col) for col in columns_to_use]

    # Create separate Excel file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'no_invoice_number_{timestamp}.xlsx'

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        report_df.to_excel(writer, sheet_name='NO_INVOICE_NUMBER', index=False)

        # Format the sheet
        from openpyxl.styles import Font, PatternFill, Alignment
        worksheet = writer.sheets['NO_INVOICE_NUMBER']

        # Header formatting
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")  # Crimson red
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Freeze top row
        worksheet.freeze_panes = 'A2'

    print(f"   🚫 Found {len(no_invoice_df)} payment advices WITHOUT invoice numbers")
    print(f"   📄 Separate report generated: {filename}")
    return filename


def main():
    """Main reconciliation workflow"""
    print("=" * 70)
    print("💰 PAYMENT RECONCILIATION SYSTEM - WARSOFT INTEGRATION")
    print("=" * 70)
    print("1. 📧 Extract payment advices from email inbox")
    print("2. 📥 Sync unpaid invoices from Warsoft")
    print("3. 🔄 Match payment advices with invoices by invoice number")
    print("4. 📤 Write matched payments to Warsoft")
    print("5. 📊 Generate Excel report (MATCHED/UNMATCHED/NOT_FOUND)")
    print("=" * 70)

    # Initialize components
    db = ReconciliationDB()
    extractor = PaymentAdviceExtractor()
    warsoft_client = WarsoftClient()
    reconciler = ReconciliationEngine(db, warsoft_client)

    # Clear old data to start fresh each run
    print("\n🗑️  Clearing previous data...")
    db.clear_payment_advices()
    db.clear_reconciliation_results()

    if not warsoft_client.enabled:
        print("\n❌ Warsoft API is not configured. Please set credentials in .env file")
        print("📋 Required: WARSOFT_ACCESS_TOKEN (or ACCESS_TOKEN)")
        return

    # Step 1: Extract payment advices from inbox
    print("\n📧 STEP 1: Extracting payment advices from inbox...")
    days_back = int(os.getenv('DAYS_TO_SEARCH', 365))
    payment_advices = extractor.fetch_payment_advices_from_email(days_back)

    if not payment_advices:
        print("⚠️  No payment advices found in inbox")
    else:
        # Store in database
        print(f"\n💾 Storing {len(payment_advices)} payment advices in database...")
        stored_count = 0
        skipped_count = 0
        for payment in payment_advices:
            try:
                result = db.insert_payment_advice(payment)
                inv_num = payment.get('invoice_number', 'Unknown')
                if result is not None:
                    # Successfully stored (not a duplicate)
                    stored_count += 1
                    print(f"   ✅ Stored: Invoice {inv_num} - ₹{payment.get('payment_amount', 0)}")
                else:
                    # Skipped duplicate (already printed by database function)
                    skipped_count += 1
            except Exception as e:
                print(f"   ⚠️  Error storing payment: {e}")

        print(f"\n📊 Storage Summary: {stored_count} stored, {skipped_count} duplicates skipped")

    # Step 2: Sync invoices from Warsoft
    print("\n📥 STEP 2: Syncing unpaid invoices from Warsoft...")
    sync_invoices_from_warsoft(db, warsoft_client)

    # Step 2.5: Load invoice cache into memory for fast reconciliation
    print("\n🚀 OPTIMIZATION: Loading invoice cache into memory...")
    cache_count = reconciler.load_invoice_cache()
    print(f"   ⚡ Ready for high-speed reconciliation with {cache_count} invoices in memory")

    # Step 3: Reconciliation (now 50-100x faster with in-memory cache!)
    print("\n🔄 STEP 3: Reconciling payments with invoices by invoice number...")
    results = reconciler.reconcile_all_pending()

    if not results:
        print("⚠️  No payments to reconcile")
        return

    # Step 4: Generate Excel report
    print("\n📊 STEP 4: Generating Excel report...")
    report_file = generate_excel_report(db)

    # Step 5: Generate separate report for payment advices without invoice numbers
    no_invoice_file = generate_no_invoice_report(db)

    # Print summary
    print("\n" + "=" * 70)
    print("✅ RECONCILIATION COMPLETE")
    print("=" * 70)

    summary = db.get_reconciliation_summary()
    total_matched = 0
    total_unmatched = 0
    total_not_found = 0

    for row in summary:
        status = row['match_status']
        count = row['count']

        if status == 'MATCHED':
            total_matched = count
            print(f"   ✅ MATCHED: {count} payments")
        elif status in ['UNMATCHED', 'PARTIAL_MATCH']:
            total_unmatched += count
            print(f"   ⚠️  UNMATCHED: {count} payments")
            if row['amount_mismatches'] > 0:
                print(f"      💵 Amount mismatches: {row['amount_mismatches']}")
            if row['total_difference']:
                print(f"      💵 Total difference: ₹{row['total_difference']:.2f}")
        elif status == 'NOT_FOUND':
            total_not_found = count
            print(f"   ❌ NOT FOUND: {count} payments (invoice missing in Zoho)")

    print("\n" + "=" * 70)
    if report_file:
        print(f"📄 Excel Report: {report_file}")
        print("   - Sheet 'MATCHED': Successfully reconciled payments")
        print("   - Sheet 'UNMATCHED': Payments with discrepancies")
        print("   - Sheet 'NOT_FOUND': Payments without matching invoices")
        print("   - Sheet 'SUMMARY': Overview statistics")
    print("=" * 70)


if __name__ == "__main__":
    main()
